import openpyxl
from analysis_oldversion import MobSF_result


class perm_result(MobSF_result):
    def analyse_perm(self):
        global index
        self.perm = {}
        for item in self.content:
            count = 0
            if item['permissions']:
                index = index + 1
                for perm in item['permissions'].keys():
                    if item['permissions'][perm]['status'] == "dangerous":
                        count = count + 1
                worksheet1.cell(index, 1, item["file_name"])
                worksheet1.cell(index, 1, self.colname)
                worksheet1.cell(index, 1, count)
                worksheet1.cell(index, 1, 1)


workbook1 = openpyxl.Workbook()
worksheet1 = workbook1.active
worksheet1.title = "permission_sheet"
# worksheet2 = workbook.create_sheet()  # 默认插在工作簿末尾
# worksheet2.title = "New Title"
# 写入第一行数据，行号和列号都从1开始计数
worksheet1.cell(1, 1, "participant")
worksheet1.cell(1, 2, "Category")
worksheet1.cell(1, 3, "Value")
worksheet1.cell(1, 4, "YLoc")
index = 1
Project = ['Analysis_name','GooglePlay','360', 'wandoujia', 'xiaomi']
for col in Project[1:]:
    a = perm_result(col)
    a.analyse_perm()



