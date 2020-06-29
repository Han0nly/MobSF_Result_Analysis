import pymongo, json
import logging
import config
import openpyxl


class MobSF_result:
    def __init__(self,colname):
        """
        MobSF_result object constructor.
        """
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
        self.dbip = config.dbip
        self.dbport = config.dbport
        if colname:
            self.colname = colname
        else:
            self.colname = config.colname
        self.dbname = config.dbname
        self._connect()

    ##############################
    # Internal Methods #
    ##############################

    def _connect(self):
        client = pymongo.MongoClient(f'mongodb://{self.dbip}:{self.dbport}')
        db = client[self.dbname]
        self.conn = db[self.colname]
        self._getcontents()

    def _getcontents(self):
        self.content = []
        self.count = 0
        for item in self.conn.find():
            self.count = self.count+1
            self.content.append(item['result'])

    ##############################
    # External Methods #
    ##############################

    def write_contents2file(self, file_name: str = 'mobsf_result.json'):
        with open(file_name, 'w+') as f:
            f.write(json.dumps(self.content, ensure_ascii=False, indent=2))


    def analyse_certificate(self):
        self.cert_result = {}
        self.cert_result['v1_false'] = 0
        self.cert_result['v2_false'] = 0
        self.cert_result['v3_false'] = 0
        self.cert_result['sha1withrsa'] = 0
        self.cert_result['md5withrsa'] = 0
        self.cert_result['goodcert'] = 0
        for item in self.content:
            if item['certificate_analysis']:
                if 'good' in item['certificate_analysis']['description']:
                    self.cert_result['goodcert'] = self.cert_result['goodcert'] + 1
                elif 'SHA1withRSA' in item['certificate_analysis']['description']:
                    self.cert_result['sha1withrsa'] = self.cert_result['sha1withrsa'] + 1
                certificate_info = item['certificate_analysis']['certificate_info']
                split_info = certificate_info.split('\n')
                for line in split_info:
                    # v1 signature:
                    split_line = line.split(':')
                    if 'v1' in split_line[0]:
                        if 'True' not in split_line[1]:
                            self.cert_result['v1_false'] = self.cert_result['v1_false'] + 1
                    # v2 signature:
                    elif 'v2' in split_line[0]:
                        if 'True' not in split_line[1]:
                            self.cert_result['v2_false'] = self.cert_result['v2_false'] + 1
                    # v3 signature:
                    elif 'v3' in split_line[0]:
                        if 'True' not in split_line[1]:
                            self.cert_result['v3_false'] = self.cert_result['v3_false'] + 1

    def analyse_permissions(self):
        self.permissions = {}
        for item in self.content:
            if item['permissions']:
                for perm in item['permissions'].keys():
                    if item['permissions'][perm]['status'] == "dangerous":
                        if perm in self.permissions.keys():
                            self.permissions[perm] = self.permissions[perm] + 1
                        else:
                            self.permissions[perm] = 1
                    else:
                        continue

    def analyse_manifest(self):
        self.manifest = {}
        self.manifest['Clear_text'] = 0
        for item in self.content:
            for weakness in item['manifest_analysis']:
                # 判断cleartest选项
                if 'Clear' == weakness['title'][:5]:
                    self.manifest['Clear_text'] = self.manifest['Clear_text'] + 1

    def binary_analysis(self):
        self.binary = {}
        self.binary['PIE'] = 0
        for item in self.content:
            for weakness in item['binary_analysis']:
                # 判断PIE选项
                if 'Position Independent Executable' in weakness['title']:
                    self.binary['PIE'] = self.binary['PIE'] + 1

    def code_analysis(self):
        self.code = {}
        for item in self.content:
            if item['code_analysis']:
                for weakness in item['code_analysis'].keys():
                    if item['code_analysis'][weakness]['level'] == "high":
                        if ('code_'+weakness) in self.code.keys():
                            self.code['code_'+weakness] = self.code['code_'+weakness] + 1
                        else:
                            self.code['code_'+weakness] = 1

    def tracker_analysis(self):
        self.trackers = {}
        self.trackers['trackers_count'] = 0
        self.trackers['trackers_300-320'] = 0
        self.trackers['trackers_321-340'] = 0
        self.trackers['trackers_341-360'] = 0
        self.trackers['trackers_361-380'] = 0
        self.trackers['trackers_381-400'] = 0
        self.trackers['trackers_401+'] = 0
        for item in self.content:
            if item['trackers'] and item['trackers']['trackers']:
                for t in item['trackers']['trackers']:
                    for identified_trackers in t.keys():
                        if ('trackers_'+identified_trackers) in self.trackers.keys():
                            self.trackers['trackers_'+identified_trackers] = self.trackers['trackers_'+identified_trackers] + 1
                        else:
                            self.trackers['trackers_'+identified_trackers] = 1
            if item['trackers'] and item['trackers']['total_trackers']:
                self.trackers['trackers_count'] = self.trackers['trackers_count'] + item['trackers']['total_trackers']
                if item['trackers']['total_trackers']<=320:
                    self.trackers['trackers_300-320'] = self.trackers['trackers_300-320'] + 1
                elif item['trackers']['total_trackers']<=340:
                    self.trackers['trackers_321-340'] = self.trackers['trackers_321-340'] + 1
                elif item['trackers']['total_trackers']<=360:
                    self.trackers['trackers_341-360'] = self.trackers['trackers_341-360'] + 1
                elif item['trackers']['total_trackers']<=380:
                    self.trackers['trackers_361-380'] = self.trackers['trackers_361-380'] + 1
                elif item['trackers']['total_trackers']<=400:
                    self.trackers['trackers_381-400'] = self.trackers['trackers_381-400'] + 1
                else:
                    self.trackers['trackers_401+'] = self.trackers['trackers_401+'] + 1

    def domains_analysis(self):
        self.domains = {}
        for item in self.content:
            if item['domains']:
                for dom in item['domains'].keys():
                    if item['domains'][dom]['bad'] == "no":
                        keyname = 'domains_good_' + dom
                    else:
                        keyname = 'domains_bad_' + dom
                    if keyname in self.domains.keys():
                        self.domains[keyname] = self.domains[keyname] + 1
                    else:
                        self.domains[keyname] = 1


    def virustotal(self):
        self.virus = 0
        for item in self.content:
            if item['virus_total']:
                if 'positives' in item['virus_total'].keys() and item['virus_total']['positives']>0:
                    self.virus=self.virus+1

    def exported_count(self):
        self.exported = {}
        self.exported['exported_activities'] = 0
        self.exported['exported_receivers'] = 0
        self.exported['exported_providers'] = 0
        self.exported['exported_services'] = 0
        for item in self.content:
            self.exported['exported_activities'] = self.exported['exported_activities'] + item['exported_count']['exported_activities']
            self.exported['exported_receivers'] = self.exported['exported_receivers'] + item['exported_count']['exported_receivers']
            self.exported['exported_providers'] = self.exported['exported_providers'] + item['exported_count']['exported_providers']
            self.exported['exported_services'] = self.exported['exported_services'] + item['exported_count']['exported_services']


    def analyse_all(self):
        self.analyse_certificate()
        self.analyse_permissions()
        self.analyse_manifest()
        self.binary_analysis()
        self.code_analysis()
        self.tracker_analysis()
        self.virustotal()
        self.domains_analysis()
        self.exported_count()


def main():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = config.colname
    # worksheet2 = workbook.create_sheet()  # 默认插在工作簿末尾
    # worksheet2.title = "New Title"
    Project = ['Analysis_name','PHOTOGRAPHY', 'ANDROID_WEAR', 'COMICS', 'PRODUCTIVITY', 'PERSONALIZATION', 'BOOKS_AND_REFERENCE', 'FINANCE', 'SPORTS', 'SOCIAL', 'PARENTING', 'LIFESTYLE', 'EVENTS', 'HOUSE_AND_HOME', 'BUSINESS', 'MAPS_AND_NAVIGATION', 'SHOPPING', 'ENTERTAINMENT', 'ART_AND_DESIGN', 'EDUCATION', 'TOOLS', 'NEWS_AND_MAGAZINES', 'WEATHER', 'LIBRARIES_AND_DEMO', 'AUTO_AND_VEHICLES', 'VIDEO_PLAYERS', 'FAMILY', 'DATING', 'HEALTH_AND_FITNESS', 'MUSIC_AND_AUDIO', 'TRAVEL_AND_LOCAL', 'BEAUTY', 'MEDICAL', 'FOOD_AND_DRINK', 'COMMUNICATION', 'GAME']
    # 写入第一行数据，行号和列号都从1开始计数
    for i in range(len(Project)):
        worksheet.cell(1, i + 1, Project[i])

    index = 0
    results_key = []
    results_content = []

    for col in Project[1:]:
        result = MobSF_result(col)
        result.analyse_all()
        if index == 0:
            # 写入第一列数据，第一行已经有数据了，i+2
            results_key = list(result.cert_result.keys()) + list(result.permissions.keys()) + list(result.binary.keys()) + list(result.trackers.keys()) + list(result.exported.keys()) + list(result.manifest.keys()) + list(result.code.keys())
            for i in range(len(results_key)):
                worksheet.cell(i + 2, 1, results_key[i])
            for key in result.cert_result.keys():
                results_content.append("{:.2f}%\n".format(100 * result.cert_result[key] / result.count))
            for key in result.permissions.keys():
                results_content.append("{:.2f}%\n".format(100 * result.permissions[key] / result.count))
            for key in result.domains.keys():
                results_content.append("{:.2f}%\n".format(100 * result.cert_result[key] / result.count))
            for key in result.binary.keys():
                results_content.append("{:.2f}%\n".format(100 * result.binary[key] / result.count))
            for key in result.trackers.keys():
                results_content.append("{:.2f}%\n".format(100 * result.trackers[key] / result.count))
            for key in result.exported.keys():
                results_content.append("{:.2f}%\n".format(100 * result.exported[key] / result.count))
            for key in result.manifest.keys():
                results_content.append("{:.2f}%\n".format(100 * result.manifest[key] / result.count))
            for key in result.code.keys():
                results_content.append("{:.2f}%\n".format(100 * result.code[key] / result.count))
            # 写入第二列数据
            for i in range(len(results_content)):
                worksheet.cell(i + 2, 2, results_content[i])
        else:
            for i in range(len(results_content)):
                results_content[i] = "0.00%"
            for key in result.cert_result.keys():
                try:
                    results_content[results_key.index(key)]="{:.2f}%\n".format(
                        100 * result.cert_result[key] / result.count)
                except ValueError:
                    results_key.append(key)
                    worksheet.cell(len(results_key) + 1, 1, key)
                    results_content.append("{:.2f}%\n".format(
                        100 * result.cert_result[key] / result.count))
            for key in result.permissions.keys():
                try:
                    results_content[results_key.index(key)]="{:.2f}%\n".format(
                        100 * result.permissions[key] / result.count)
                except ValueError:
                    results_key.append(key)
                    worksheet.cell(len(results_key) + 1, 1, key)
                    results_content.append("{:.2f}%\n".format(
                        100 * result.permissions[key] / result.count))
            for key in result.domains.keys():
                try:
                    results_content[results_key.index(key)] = "{:.2f}%\n".format(
                        100 * result.domains[key] / result.count)
                except ValueError:
                    results_key.append(key)
                    worksheet.cell(len(results_key) + 1, 1, key)
                    results_content.append("{:.2f}%\n".format(100 * result.domains[key] / result.count))
            for key in result.binary.keys():
                try:
                    results_content[results_key.index(key)]="{:.2f}%\n".format(
                        100 * result.binary[key] / result.count)
                except ValueError:
                    results_key.append(key)
                    worksheet.cell(len(results_key) + 1, 1, key)
                    results_content.append("{:.2f}%\n".format(100 * result.binary[key] / result.count))
            for key in result.trackers.keys():
                try:
                    results_content[results_key.index(key)]="{:.2f}%\n".format(
                        100 * result.trackers[key] / result.count)
                except ValueError:
                    results_key.append(key)
                    worksheet.cell(len(results_key) + 1, 1, key)
                    results_content.append("{:.2f}%\n".format(100 * result.trackers[key] / result.count))
            for key in result.exported.keys():
                try:
                    results_content[results_key.index(key)]="{:.2f}%\n".format(
                        100 * result.exported[key] / result.count)
                except ValueError:
                    results_key.append(key)
                    worksheet.cell(len(results_key) + 1, 1, key)
                    results_content.append("{:.2f}%\n".format(100 * result.exported[key] / result.count))
            for key in result.manifest.keys():
                try:
                    results_content[results_key.index(key)]="{:.2f}%\n".format(
                        100 * result.manifest[key] / result.count)
                except ValueError:
                    results_key.append(key)
                    worksheet.cell(len(results_key) + 1, 1, key)
                    results_content.append("{:.2f}%\n".format(100 * result.manifest[key] / result.count))
            for key in result.code.keys():
                try:
                    results_content[results_key.index(key)] = "{:.2f}%\n".format(
                        100 * result.code[key] / result.count)
                except ValueError:
                    results_key.append(key)
                    worksheet.cell(len(results_key) + 1, 1, key)
                    results_content.append("{:.2f}%\n".format(100 * result.code[key] / result.count))
            for i in range(len(results_content)):
                worksheet.cell(i + 2, index + 2, results_content[i])
        index = index+1

    workbook.save(filename='./result.xlsx')


if __name__ == '__main__':
    main()
